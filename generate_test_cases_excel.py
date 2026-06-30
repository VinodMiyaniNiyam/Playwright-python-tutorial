import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Define detailed test cases data for each of the 20 functions
test_cases = [
    # 1. Student Grade Calculator
    {"fn": "calculate_grade", "id": "TC_01_01", "desc": "Score of 90 and above gets A", "inputs": "95", "expected": "A", "explanation": "95 is >= 90, so it maps to grade A"},
    {"fn": "calculate_grade", "id": "TC_01_02", "desc": "Score of 75-89 gets B", "inputs": "82", "expected": "B", "explanation": "82 is between 75 and 89, so it maps to grade B"},
    {"fn": "calculate_grade", "id": "TC_01_03", "desc": "Score of 60-74 gets C", "inputs": "67", "expected": "C", "explanation": "67 is between 60 and 74, so it maps to grade C"},
    {"fn": "calculate_grade", "id": "TC_01_04", "desc": "Score of 40-59 gets D", "inputs": "50", "expected": "D", "explanation": "50 is between 40 and 59, so it maps to grade D"},
    {"fn": "calculate_grade", "id": "TC_01_05", "desc": "Score below 40 gets F", "inputs": "35", "expected": "F", "explanation": "35 is less than 40, so it maps to grade F"},
    {"fn": "calculate_grade", "id": "TC_01_06", "desc": "Score out of range (upper bound) returns Invalid", "inputs": "105", "expected": "Invalid", "explanation": "105 is > 100, which is outside the range [0, 100]"},
    {"fn": "calculate_grade", "id": "TC_01_07", "desc": "Score out of range (lower bound) returns Invalid", "inputs": "-5", "expected": "Invalid", "explanation": "-5 is < 0, which is outside the range [0, 100]"},
    {"fn": "calculate_grade", "id": "TC_01_08", "desc": "Non-numeric input returns Invalid", "inputs": "'abc'", "expected": "Invalid", "explanation": "Input is a string, which is non-numeric"},

    # 2. Electricity Bill Calculator
    {"fn": "electricity_bill", "id": "TC_02_01", "desc": "Units <= 100 (5 per unit)", "inputs": "50", "expected": "250.0", "explanation": "50 * 5.0 = 250.0"},
    {"fn": "electricity_bill", "id": "TC_02_02", "desc": "Units <= 200 (100 * 5 + next * 7)", "inputs": "150", "expected": "850.0", "explanation": "(100 * 5.0) + (50 * 7.0) = 500.0 + 350.0 = 850.0"},
    {"fn": "electricity_bill", "id": "TC_02_03", "desc": "Units > 200 (100 * 5 + 100 * 7 + next * 10)", "inputs": "250", "expected": "1700.0", "explanation": "(100 * 5.0) + (100 * 7.0) + (50 * 10.0) = 500 + 700 + 500 = 1700.0. Note: PDF example shows 1650, but text slab rules compute to 1700.0"},
    {"fn": "electricity_bill", "id": "TC_02_04", "desc": "Negative units returns 0", "inputs": "-10", "expected": "0.0", "explanation": "Units cannot be negative"},

    # 3. Simple ATM Withdrawal
    {"fn": "atm_withdraw", "id": "TC_03_01", "desc": "Normal withdrawal with valid amount", "inputs": "balance=5000, amount=1200", "expected": "3800", "explanation": "Withdrawal is possible. New balance = 5000 - 1200 = 3800"},
    {"fn": "atm_withdraw", "id": "TC_03_02", "desc": "Withdrawal with negative amount", "inputs": "balance=5000, amount=-100", "expected": "'Amount must be positive'", "explanation": "Amount must be greater than zero"},
    {"fn": "atm_withdraw", "id": "TC_03_03", "desc": "Withdrawal with amount not multiple of 100", "inputs": "balance=5000, amount=1250", "expected": "'Amount must be a multiple of 100'", "explanation": "ATM only dispenses 100 bills, so 1250 is rejected"},
    {"fn": "atm_withdraw", "id": "TC_03_04", "desc": "Withdrawal exceeding available balance", "inputs": "balance=1000, amount=1500", "expected": "'Insufficient balance'", "explanation": "1500 > 1000, which is disallowed"},

    # 4. Count Vowels and Consonants
    {"fn": "count_letters", "id": "TC_04_01", "desc": "Counts standard string with spaces", "inputs": "'Python Class'", "expected": "{'vowels': 2, 'consonants': 9}", "explanation": "Letters: P,y,t,h,o,n,C,l,a,s,s. Vowels: o, a (2). Consonants: P,y,t,h,n,C,l,s,s (9)"},
    {"fn": "count_letters", "id": "TC_04_02", "desc": "Ignores numbers and special characters", "inputs": "'A1b!'", "expected": "{'vowels': 1, 'consonants': 1}", "explanation": "Vowels: A (1). Consonants: b (1). 1 and ! are ignored"},

    # 5. Remove Duplicate Characters
    {"fn": "remove_duplicates", "id": "TC_05_01", "desc": "Removes duplicate chars in lowercase string", "inputs": "'programming'", "expected": "'progamin'", "explanation": "Only first occurrences kept: p,r,o,g,a,m,i,n"},
    {"fn": "remove_duplicates", "id": "TC_05_02", "desc": "Case-sensitive duplicate removal with spaces", "inputs": "'Hello World'", "expected": "'Helo Wrd'", "explanation": "Keep first: H,e,l,o,' ',W,r,d (second l, second o, second l are duplicates)"},

    # 6. Check Strong Password
    {"fn": "is_strong_password", "id": "TC_06_01", "desc": "Valid strong password", "inputs": "'Abc@1234'", "expected": "True", "explanation": "Length 8, contains 'A' (upper), 'b' (lower), '1' (digit), '@' (special)"},
    {"fn": "is_strong_password", "id": "TC_06_02", "desc": "Weak password lacking uppercase and special char", "inputs": "'password123'", "expected": "False", "explanation": "Missing uppercase letter and special character"},
    {"fn": "is_strong_password", "id": "TC_06_03", "desc": "Weak password under 8 characters", "inputs": "'A@1bc'", "expected": "False", "explanation": "Meets criteria but length is 5 (< 8)"},

    # 7. Second Largest Number in a List
    {"fn": "second_largest", "id": "TC_07_01", "desc": "Finds second largest with duplicate maximums", "inputs": "[10, 40, 20, 40, 30]", "expected": "30", "explanation": "Largest is 40. Second largest strictly less than 40 is 30"},
    {"fn": "second_largest", "id": "TC_07_02", "desc": "List with all identical values", "inputs": "[5, 5, 5]", "expected": "None", "explanation": "No value exists that is strictly less than the maximum (5)"},
    {"fn": "second_largest", "id": "TC_07_03", "desc": "List with fewer than 2 elements", "inputs": "[10]", "expected": "None", "explanation": "Need at least 2 numbers to have a second largest"},

    # 8. Count Positive, Negative and Zero Values
    {"fn": "count_numbers", "id": "TC_08_01", "desc": "Counts pos, neg and zeros in list", "inputs": "[5, -2, 0, 7, -9, 0]", "expected": "{'positive': 2, 'negative': 2, 'zero': 2}", "explanation": "Positives: 5, 7. Negatives: -2, -9. Zeros: 0, 0"},
    {"fn": "count_numbers", "id": "TC_08_02", "desc": "Empty list", "inputs": "[]", "expected": "{'positive': 0, 'negative': 0, 'zero': 0}", "explanation": "All counters remain at 0"},

    # 9. List Average Without sum()
    {"fn": "average", "id": "TC_09_01", "desc": "Calculates average of list", "inputs": "[10, 20, 30, 40]", "expected": "25.0", "explanation": "(10 + 20 + 30 + 40) / 4 = 100 / 4 = 25.0"},
    {"fn": "average", "id": "TC_09_02", "desc": "Empty list returns Invalid", "inputs": "[]", "expected": "'Invalid'", "explanation": "Cannot divide by zero for empty list"},

    # 10. Search Element Manually
    {"fn": "search_item", "id": "TC_10_01", "desc": "Item exists at multiple indices (first returned)", "inputs": "numbers=[4, 8, 1, 8], target=8", "expected": "1", "explanation": "First occurrence of 8 is at index 1"},
    {"fn": "search_item", "id": "TC_10_02", "desc": "Item not found in list", "inputs": "numbers=[4, 8, 1, 8], target=10", "expected": "-1", "explanation": "10 does not exist in the list"},

    # 11. Manual String Length
    {"fn": "string_length", "id": "TC_11_01", "desc": "Measures length of standard string", "inputs": "'Python'", "expected": "6", "explanation": "Counts 6 characters"},
    {"fn": "string_length", "id": "TC_11_02", "desc": "Measures empty string", "inputs": "''", "expected": "0", "explanation": "Counts 0 characters"},

    # 12. Reverse a String Manually
    {"fn": "reverse_text", "id": "TC_12_01", "desc": "Reverses standard string", "inputs": "'function'", "expected": "'noitcnuf'", "explanation": "Characters reversed character-by-character"},
    {"fn": "reverse_text", "id": "TC_12_02", "desc": "Reverses single character", "inputs": "'a'", "expected": "'a'", "explanation": "Reversing 'a' is 'a'"},

    # 13. Check Palindrome String
    {"fn": "is_text_palindrome", "id": "TC_13_01", "desc": "Palindrome ignoring space and case", "inputs": "'Never odd or even'", "expected": "True", "explanation": "Cleaned string is 'neveroddoreven', which reads same forwards and backwards"},
    {"fn": "is_text_palindrome", "id": "TC_13_02", "desc": "Non-palindrome string", "inputs": "'hello'", "expected": "False", "explanation": "'hello' != 'olleh'"},

    # 14. Print Right Triangle Pattern
    {"fn": "print_triangle", "id": "TC_14_01", "desc": "Triangle pattern of 4 rows", "inputs": "4", "expected": "'*\\n**\\n***\\n****'", "explanation": "Prints 4 lines with 1, 2, 3, and 4 stars"},
    {"fn": "print_triangle", "id": "TC_14_02", "desc": "Triangle pattern of 1 row", "inputs": "1", "expected": "'*'", "explanation": "Prints 1 line with 1 star"},

    # 15. Find All Factors
    {"fn": "factors", "id": "TC_15_01", "desc": "Factors of composite number", "inputs": "12", "expected": "[1, 2, 3, 4, 6, 12]", "explanation": "Factors of 12 are numbers that divide 12 fully: 1,2,3,4,6,12"},
    {"fn": "factors", "id": "TC_15_02", "desc": "Factors of prime number", "inputs": "7", "expected": "[1, 7]", "explanation": "7 is prime, only divisible by 1 and 7"},

    # 16. Check Prime Using Function
    {"fn": "is_prime", "id": "TC_16_01", "desc": "Checks prime number", "inputs": "29", "expected": "True", "explanation": "29 has no positive divisors other than 1 and itself"},
    {"fn": "is_prime", "id": "TC_16_02", "desc": "Checks composite number", "inputs": "4", "expected": "False", "explanation": "4 is divisible by 2"},
    {"fn": "is_prime", "id": "TC_16_03", "desc": "Checks base case 1", "inputs": "1", "expected": "False", "explanation": "1 is not prime by definition"},

    # 17. Print Prime Numbers in a Range
    {"fn": "primes_between", "id": "TC_17_01", "desc": "Primes in range with multiple results", "inputs": "start=10, end=20", "expected": "[11, 13, 17, 19]", "explanation": "Primes between 10 and 20 are 11, 13, 17, 19"},
    {"fn": "primes_between", "id": "TC_17_02", "desc": "Primes in range with zero results", "inputs": "start=8, end=10", "expected": "[]", "explanation": "No primes exist between 8 and 10"},

    # 18. Simple Calculator Function
    {"fn": "calculator", "id": "TC_18_01", "desc": "Modulo operation", "inputs": "a=20, b=6, operator='%'", "expected": "2", "explanation": "20 % 6 = 2 (remainder of 20 / 6)"},
    {"fn": "calculator", "id": "TC_18_02", "desc": "Division by zero", "inputs": "a=5, b=0, operator='/'", "expected": "'Error: Division by zero'", "explanation": "Cannot divide by 0, returns error message"},
    {"fn": "calculator", "id": "TC_18_03", "desc": "Invalid operator", "inputs": "a=5, b=5, operator='?'", "expected": "'Error: Invalid operator'", "explanation": "? is not a supported operator"},

    # 19. Employee Salary Calculation
    {"fn": "net_salary", "id": "TC_19_01", "desc": "Net salary for basic 20000", "inputs": "20000", "expected": "24700.0", "explanation": "HRA=4000, DA=2000, Gross=26000, Tax(5%)=1300. Net = 26000 - 1300 = 24700.0"},
    {"fn": "net_salary", "id": "TC_19_02", "desc": "Net salary for basic 0", "inputs": "0", "expected": "0.0", "explanation": "Basic 0 results in 0 net salary"},

    # 20. Shopping Cart Total
    {"fn": "cart_total", "id": "TC_20_01", "desc": "Total with discount (> 5000)", "inputs": "prices=[1000, 500, 200], quantities=[3, 4, 5]", "expected": "5400.0", "explanation": "Subtotal = 3000+2000+1000 = 6000. Apply 10% discount: 6000 * 0.90 = 5400.0"},
    {"fn": "cart_total", "id": "TC_20_02", "desc": "Total without discount (<= 5000)", "inputs": "prices=[100, 200], quantities=[2, 3]", "expected": "800.0", "explanation": "Subtotal = 200 + 600 = 800. <= 5000, no discount applied"},

    # --- Optional Challenge (Marks Management System) ---
    {"fn": "add_student", "id": "TC_OC_01", "desc": "Add student record", "inputs": "names=[], marks=[], name='Alice', mark=85", "expected": "True", "explanation": "Student name and mark are successfully appended"},
    {"fn": "show_students", "id": "TC_OC_02", "desc": "Format student list", "inputs": "names=['Alice', 'Bob'], marks=[85, 92]", "expected": "'Student Records:\\n- Alice: 85\\n- Bob: 92'", "explanation": "Formats names and marks in a human-readable string"},
    {"fn": "find_highest_marks", "id": "TC_OC_03", "desc": "Find highest student mark", "inputs": "names=['Alice', 'Bob'], marks=[85, 92]", "expected": "('Bob', 92)", "explanation": "Bob has the maximum mark of 92"},
    {"fn": "calculate_class_average", "id": "TC_OC_04", "desc": "Calculate marks average", "inputs": "marks=[85, 92, 78]", "expected": "85.0", "explanation": "(85 + 92 + 78) / 3 = 255 / 3 = 85.0"},
    {"fn": "search_student", "id": "TC_OC_05", "desc": "Search existing student", "inputs": "names=['Alice', 'Bob'], marks=[85, 92], name='Bob'", "expected": "1", "explanation": "Bob is found at index 1"},
    {"fn": "search_student", "id": "TC_OC_06", "desc": "Search non-existing student", "inputs": "names=['Alice', 'Bob'], name='Charlie'", "expected": "-1", "explanation": "Charlie is not in list, returns -1"},
]

# Create workbook and sheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Test Cases & Test Data"

# Enable grid lines explicitly
ws.views.sheetView[0].showGridLines = True

# Style setup
header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid") # Dark Blue
data_font = Font(name="Segoe UI", size=10)
center_align = Alignment(horizontal="center", vertical="center")
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

thin_side = Side(border_style="thin", color="D3D3D3")
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

# Write headers
headers = ["Function Name", "Test Case ID", "Description", "Test Inputs (Arguments)", "Expected Output", "Test Data Explanation"]
ws.append(headers)

# Style headers
for col_idx in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align if col_idx in [2, 5] else left_align
    cell.border = thin_border
ws.row_dimensions[1].height = 25

# Write data rows
for tc in test_cases:
    row_data = [tc["fn"], tc["id"], tc["desc"], tc["inputs"], tc["expected"], tc["explanation"]]
    ws.append(row_data)
    
    current_row = ws.max_row
    ws.row_dimensions[current_row].height = 24
    
    # Style standard cells
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.font = data_font
        cell.border = thin_border
        
        # Specific alignments
        if col_idx == 2: # TC ID
            cell.alignment = center_align
        else:
            cell.alignment = left_align

# Column widths
column_widths = {
    "A": 22, # Function Name
    "B": 12, # TC ID
    "C": 35, # Description
    "D": 38, # Inputs
    "E": 25, # Expected Output
    "F": 55  # Explanation
}

for col_letter, width in column_widths.items():
    ws.column_dimensions[col_letter].width = width

# Save the workbook
output_path = "Practical_Assignment_Test_Cases.xlsx"
wb.save(output_path)
print(f"Excel test cases sheet created successfully at: {output_path}")
