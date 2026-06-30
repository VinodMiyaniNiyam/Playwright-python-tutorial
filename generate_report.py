import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Define the package data
packages = [
    {
        "name": "Python 3.11.9",
        "scope": "System (via winget)",
        "version": "3.11.9",
        "install_time": "2026-06-30 11:59:11",
        "link": "https://www.python.org/",
        "description": "Core programming language runtime"
    },
    {
        "name": "pip",
        "scope": "Virtual Environment",
        "version": "26.1.2",
        "install_time": "2026-06-30 12:00:32",
        "link": "https://pip.pypa.io/",
        "description": "Package installer for Python"
    },
    {
        "name": "playwright",
        "scope": "Virtual Environment",
        "version": "1.61.0",
        "install_time": "2026-06-30 12:08:41",
        "link": "https://playwright.dev/python/",
        "description": "Core library for browser automation"
    },
    {
        "name": "pytest-playwright",
        "scope": "Virtual Environment",
        "version": "0.8.0",
        "install_time": "2026-06-30 12:08:41",
        "link": "https://github.com/microsoft/playwright-pytest",
        "description": "Pytest plugin for writing end-to-end tests with Playwright"
    },
    {
        "name": "pytest",
        "scope": "Virtual Environment",
        "version": "9.1.1",
        "install_time": "2026-06-30 12:08:41",
        "link": "https://docs.pytest.org/",
        "description": "Testing framework for writing and running test cases"
    },
    {
        "name": "pytest-xdist",
        "scope": "Virtual Environment",
        "version": "3.8.0",
        "install_time": "2026-06-30 12:21:34",
        "link": "https://github.com/pytest-dev/pytest-xdist",
        "description": "Plugin for running tests in parallel across multiple CPUs"
    },
    {
        "name": "pytest-forked",
        "scope": "Virtual Environment",
        "version": "1.6.0",
        "install_time": "2026-06-30 12:21:34",
        "link": "https://github.com/pytest-dev/pytest-forked",
        "description": "Plugin for running each test in a forked subprocess"
    },
    {
        "name": "execnet",
        "scope": "Virtual Environment",
        "version": "2.1.2",
        "install_time": "2026-06-30 12:21:34",
        "link": "https://github.com/pytest-dev/execnet",
        "description": "Ad-hoc multi-interpreter creation and communication library"
    },
    {
        "name": "toml",
        "scope": "Virtual Environment",
        "version": "0.10.2",
        "install_time": "2026-06-30 12:21:34",
        "link": "https://github.com/uiri/toml",
        "description": "Python library for parsing TOML files"
    },
    {
        "name": "tomli",
        "scope": "Virtual Environment",
        "version": "2.4.1",
        "install_time": "2026-06-30 12:21:34",
        "link": "https://github.com/hukkin/tomli",
        "description": "Fast little TOML parser for Python"
    },
    {
        "name": "greenlet",
        "scope": "Virtual Environment",
        "version": "3.5.3",
        "install_time": "2026-06-30 12:08:03",
        "link": "https://github.com/python-greenlet/greenlet",
        "description": "Lightweight coroutines for in-process concurrent programming"
    },
    {
        "name": "requests",
        "scope": "Virtual Environment",
        "version": "2.34.2",
        "install_time": "2026-06-30 12:08:41",
        "link": "https://requests.readthedocs.io/",
        "description": "Elegant and simple HTTP library for Python"
    },
    {
        "name": "urllib3",
        "scope": "Virtual Environment",
        "version": "2.7.0",
        "install_time": "2026-06-30 12:08:41",
        "link": "https://urllib3.readthedocs.io/",
        "description": "HTTP client with thread-safe connection pooling"
    },
    {
        "name": "certifi",
        "scope": "Virtual Environment",
        "version": "2026.6.17",
        "install_time": "2026-06-30 12:08:41",
        "link": "https://github.com/certifi/python-certifi",
        "description": "Python package providing Mozilla's curated CA Bundle"
    },
    {
        "name": "charset-normalizer",
        "scope": "Virtual Environment",
        "version": "3.4.7",
        "install_time": "2026-06-30 12:08:41",
        "link": "https://github.com/Ousret/charset_normalizer",
        "description": "Charset detector library for Python"
    },
    {
        "name": "idna",
        "scope": "Virtual Environment",
        "version": "3.18",
        "install_time": "2026-06-30 12:08:41",
        "link": "https://github.com/kjd/idna",
        "description": "Support for Internationalized Domain Names in Applications"
    },
    {
        "name": "colorama",
        "scope": "Virtual Environment",
        "version": "0.4.6",
        "install_time": "2026-06-30 12:08:41",
        "link": "https://github.com/tartley/colorama",
        "description": "Cross-platform colored terminal text in Python"
    },
    {
        "name": "iniconfig",
        "scope": "Virtual Environment",
        "version": "2.3.0",
        "install_time": "2026-06-30 12:08:41",
        "link": "https://github.com/pytest-dev/iniconfig",
        "description": "Brain-dead simple INI parser for Python"
    },
    {
        "name": "packaging",
        "scope": "Virtual Environment",
        "version": "26.2",
        "install_time": "2026-06-30 12:08:41",
        "link": "https://github.com/pypa/packaging",
        "description": "Core utilities for Python packages"
    },
    {
        "name": "pygments",
        "scope": "Virtual Environment",
        "version": "2.20.0",
        "install_time": "2026-06-30 12:08:41",
        "link": "https://pygments.org/",
        "description": "Syntax highlighting package written in Python"
    },
    {
        "name": "pytest-base-url",
        "scope": "Virtual Environment",
        "version": "2.1.0",
        "install_time": "2026-06-30 12:08:41",
        "link": "https://github.com/pytest-dev/pytest-base-url",
        "description": "Pytest fixture for providing base URLs for web tests"
    },
    {
        "name": "python-slugify",
        "scope": "Virtual Environment",
        "version": "8.0.4",
        "install_time": "2026-06-30 12:08:41",
        "link": "https://github.com/un33k/python-slugify",
        "description": "Python slugify library that translates Unicode to ASCII"
    },
    {
        "name": "text-unidecode",
        "scope": "Virtual Environment",
        "version": "1.3",
        "install_time": "2026-06-30 12:08:41",
        "link": "https://github.com/kmike/text-unidecode",
        "description": "Fast and lightweight ASCII transliterations of Unicode"
    },
    {
        "name": "typing-extensions",
        "scope": "Virtual Environment",
        "version": "4.15.0",
        "install_time": "2026-06-30 12:08:41",
        "link": "https://github.com/python/typing_extensions",
        "description": "Backported and experimental type hints for Python"
    },
    {
        "name": "openpyxl",
        "scope": "Virtual Environment",
        "version": "3.1.5",
        "install_time": "2026-06-30 12:26:11",
        "link": "https://openpyxl.readthedocs.io/",
        "description": "Library to read/write Excel 2010 xlsx/xlsm/xltx/xltm files"
    },
    {
        "name": "et-xmlfile",
        "scope": "Virtual Environment",
        "version": "2.0.0",
        "install_time": "2026-06-30 12:26:11",
        "link": "https://github.com/wg/et_xmlfile",
        "description": "Low memory library for creating large XML files"
    }
]

# Create workbook and sheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Installed Packages"

# Enable grid lines explicitly
ws.views.sheetView[0].showGridLines = True

# Style setup
header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid") # Classic Excel Blue
data_font = Font(name="Segoe UI", size=10)
link_font = Font(name="Segoe UI", size=10, color="0563C1", underline="single") # Standard hyperlink color
center_align = Alignment(horizontal="center", vertical="center")
left_align = Alignment(horizontal="left", vertical="center")

thin_side = Side(border_style="thin", color="D3D3D3")
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

# Write headers
headers = ["Package / Tool Name", "Installation Scope", "Installed Version", "Installation Date & Time", "Project Website / Link", "Description & Purpose"]
ws.append(headers)

# Style headers
for col_idx in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = left_align if col_idx != 3 else center_align
    cell.border = thin_border
ws.row_dimensions[1].height = 25

# Write data rows
for p in packages:
    row_data = [p["name"], p["scope"], p["version"], p["install_time"], p["link"], p["description"]]
    ws.append(row_data)
    
    current_row = ws.max_row
    ws.row_dimensions[current_row].height = 20
    
    # Style standard cells
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.font = data_font
        cell.border = thin_border
        
        # Specific alignments
        if col_idx in [3, 4]: # Version and Date Time
            cell.alignment = center_align
        else:
            cell.alignment = left_align
            
        # Formatted hyperlink for Project Link
        if col_idx == 5:
            cell.font = link_font
            cell.hyperlink = p["link"]

# Auto-fit columns with safety padding
for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            max_len = max(max_len, len(str(cell.value)))
    # Set width
    ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

# Save the workbook
output_path = "Installed_Dependencies_Report.xlsx"
wb.save(output_path)
print(f"Excel report created successfully at: {output_path}")
